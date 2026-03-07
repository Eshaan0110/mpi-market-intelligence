"""
src/ingestion.py
RBI + NPCI Data Ingestion Pipeline
====================================
Dataset 1: RBI Payment System Indicators (Table 45)
  - Monthly credit/debit card transaction volumes and values
  - Cards outstanding (credit + debit)
  - Source: RBI DBIE → Statistics → Financial Sector → Payment Systems

Dataset 2: NPCI UPI Product Statistics
  - Monthly UPI transaction volume and value
  - Number of banks live on UPI
  - Source: https://www.npci.org.in/what-we-do/upi/upi-ecosystem-statistics

Column mapping for RBI Excel (0-indexed, verified against actual file):
  1   = Month/Year
  52  = Credit Cards Volume (Lakh)
  53  = Credit Cards Value (Rupees Crores)
  58  = Debit Cards Volume (Lakh)
  59  = Debit Cards Value (Rupees Crores)
  118 = Credit Cards Outstanding (Lakh)
  120 = Debit Cards Outstanding (Lakh)
"""

import pandas as pd
import openpyxl
from pathlib import Path
from loguru import logger

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT_DIR = Path(__file__).resolve().parent.parent
RAW_DIR  = ROOT_DIR / "data" / "raw"
PROC_DIR = ROOT_DIR / "data" / "processed"
RAW_DIR.mkdir(parents=True, exist_ok=True)
PROC_DIR.mkdir(parents=True, exist_ok=True)

# ── RBI Column indices (0-based, verified) ─────────────────────────────────
COL_DATE             = 1
COL_CREDIT_VOL       = 52
COL_CREDIT_VAL       = 53
COL_DEBIT_VOL        = 58
COL_DEBIT_VAL        = 59
COL_CREDIT_CARDS_OUT = 118
COL_DEBIT_CARDS_OUT  = 120
DATA_START_ROW       = 7    # first row index of actual monthly data


# ── Shared helper ──────────────────────────────────────────────────────────

def safe_float(row, idx):
    """Return float from row[idx], or None if missing/dash/invalid.
    Handles comma-formatted numbers e.g. '21,703.44' from newer NPCI files.
    """
    val = row[idx] if len(row) > idx else None
    if val is None or str(val).strip() in ("-", ""):
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# ══════════════════════════════════════════════════════════════════════════════
# DATASET 1 — RBI Payment System Indicators
# ══════════════════════════════════════════════════════════════════════════════

def parse_psi_excel(filepath: str | Path) -> pd.DataFrame:
    """
    Parse RBI Table 45 Payment System Indicators Excel.
    Returns a clean monthly DataFrame.
    """
    filepath = Path(filepath)
    logger.info(f"Parsing: {filepath.name}")

    wb   = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[DATA_START_ROW:]:
        date_val = row[COL_DATE] if len(row) > COL_DATE else None

        # Stop at footnote rows at the bottom of the file
        if not date_val or not isinstance(date_val, str):
            continue
        try:
            date = pd.to_datetime(date_val, format="%b-%Y")
        except Exception:
            logger.debug(f"Skipping non-date row: {date_val}")
            continue

        records.append({
            "date":                          date,
            "credit_card_vol_lakh":          safe_float(row, COL_CREDIT_VOL),
            "credit_card_val_cr":            safe_float(row, COL_CREDIT_VAL),
            "debit_card_vol_lakh":           safe_float(row, COL_DEBIT_VOL),
            "debit_card_val_cr":             safe_float(row, COL_DEBIT_VAL),
            "credit_cards_outstanding_lakh": safe_float(row, COL_CREDIT_CARDS_OUT),
            "debit_cards_outstanding_lakh":  safe_float(row, COL_DEBIT_CARDS_OUT),
        })

    df = (pd.DataFrame(records)
            .sort_values("date")
            .reset_index(drop=True))

    logger.info(
        f"RBI PSI parsed: {len(df)} rows | "
        f"{df['date'].min().strftime('%b %Y')} → "
        f"{df['date'].max().strftime('%b %Y')}"
    )
    return df


def run_ingestion(excel_path: str | Path = None) -> pd.DataFrame:
    """Main entry point for RBI PSI — parse, validate, save."""
    if excel_path is None:
        matches = list(RAW_DIR.glob("*Payment_System*"))
        if not matches:
            raise FileNotFoundError(
                f"No Payment System Indicators Excel found in {RAW_DIR}\n"
                "Download from: RBI DBIE → Statistics → Financial Sector → Payment Systems"
            )
        excel_path = matches[0]
        logger.info(f"Auto-found: {excel_path.name}")

    df = parse_psi_excel(excel_path)

    # Warn on high nulls
    for col, pct in (df.isnull().mean() * 100).items():
        if pct > 20:
            logger.warning(f"High nulls in '{col}': {pct:.1f}%")

    df.to_csv(PROC_DIR / "rbi_psi_cards.csv", index=False)
    df.to_parquet(PROC_DIR / "rbi_psi_cards.parquet", index=False)
    logger.info(f"RBI PSI saved to {PROC_DIR}")

    print("\n=== RBI PSI — last 5 months ===")
    print(df.tail().to_string(index=False))
    print(f"\nTotal rows : {len(df)}")
    print(f"Date range : {df['date'].min().strftime('%b %Y')} → {df['date'].max().strftime('%b %Y')}")
    print(f"\nNull counts:\n{df.isnull().sum()}")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# DATASET 2 — NPCI UPI Product Statistics
# ══════════════════════════════════════════════════════════════════════════════

def parse_npci_upi_excel(filepath: str | Path) -> pd.DataFrame:
    """
    Parse a single NPCI UPI Product Statistics yearly Excel file.
    Format: Row 0 = header, data from row 1. Clean structure, no merged cells.

    Columns: Month | No. of Banks live on UPI | Volume (In Mn.) | Value (In Cr.)
    """
    filepath = Path(filepath)
    wb   = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[1:]:    # skip header row
        if not row[0]:
            continue
        try:
            date = pd.to_datetime(row[0], format="%B-%Y")
        except Exception:
            logger.debug(f"Skipping row: {row[0]}")
            continue

        records.append({
            "date":           date,
            "upi_banks_live": safe_float(row, 1),
            "upi_volume_mn":  safe_float(row, 2),
            "upi_value_cr":   safe_float(row, 3),
        })

    return pd.DataFrame(records)


def run_npci_ingestion() -> pd.DataFrame:
    """
    Loop through all NPCI yearly Excel files in data/raw/,
    parse each one, stack into a single DataFrame, and save.
    """
    files = sorted(RAW_DIR.glob("Product-Statistics-UPI*"))
    if not files:
        raise FileNotFoundError(
            f"No NPCI UPI Excel files found in {RAW_DIR}\n"
            "Download from: https://www.npci.org.in/what-we-do/upi/upi-ecosystem-statistics"
        )

    logger.info(f"Found {len(files)} NPCI yearly files")

    frames = []
    for f in files:
        logger.info(f"Parsing: {f.name}")
        df = parse_npci_upi_excel(f)
        frames.append(df)

    # Stack all years, sort, deduplicate
    combined = (pd.concat(frames, ignore_index=True)
                  .sort_values("date")
                  .drop_duplicates(subset=["date"])
                  .reset_index(drop=True))

    logger.info(
        f"NPCI UPI parsed: {len(combined)} months | "
        f"{combined['date'].min().strftime('%b %Y')} → "
        f"{combined['date'].max().strftime('%b %Y')}"
    )

    # Warn on high nulls
    for col, pct in (combined.isnull().mean() * 100).items():
        if pct > 10:
            logger.warning(f"High nulls in '{col}': {pct:.1f}%")

    combined.to_csv(PROC_DIR / "npci_upi.csv", index=False)
    combined.to_parquet(PROC_DIR / "npci_upi.parquet", index=False)
    logger.info(f"NPCI UPI saved to {PROC_DIR}")

    print("\n=== NPCI UPI — last 5 months ===")
    print(combined.tail().to_string(index=False))
    print(f"\nTotal rows : {len(combined)}")
    print(f"Date range : {combined['date'].min().strftime('%b %Y')} → {combined['date'].max().strftime('%b %Y')}")

    return combined


# ══════════════════════════════════════════════════════════════════════════════
# MAIN — runs both datasets
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n" + "="*50)
    print("DATASET 1 — RBI Payment System Indicators")
    print("="*50)
    run_ingestion()

    print("\n" + "="*50)
    print("DATASET 2 — NPCI UPI Product Statistics")
    print("="*50)
    run_npci_ingestion()